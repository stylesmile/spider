# !/usr/bin/env python3
# -*- coding: utf-8 -*-

# @Time    : 2019/12/19 16:08
# @File    : 多语言项目质检文本导出.py
# @Date    : 2019/12/19
# @Author  : Yuwenjun
# @Desc    : 多语言项目文本导出


import json
import codecs
import pathlib
import math
import time
import os
from pymongo import MongoClient
from openpyxl import Workbook


client = MongoClient("39.97.250.105", 27015)
mongodb = client.data
# mongodb.authenticate('tasks', 'admin123456!')


def get_col_data(col_name, count):
    col = mongodb[col_name]
    ret = list(col.find({"handleStatus": 0}).sort([("order", 1)]).limit(count))
    return ret


def switch_order(length, order):
    return {
        1: lambda x: "000" + str(order),
        2: lambda x: "00" + str(order),
        3: lambda x: "0" + str(order),
        4: lambda x: str(order)
    }[length](order)


def write_2_txt(datas, output_path):
    s, e = datas[0].get("order"), datas[-1].get("order")
    with codecs.open(output_path, "w", encoding="utf-8") as f:
        for data in datas:
            content = data.get("content").replace("\n", " ").replace("\r\n", " ")
            f.write(content + "\n")
    return s, e


def write_2_json(datas, output_path):
    path = pathlib.Path(output_path)
    json_dic = dict()
    json_dic['taskName'] = path.name.split(".")[0]
    json_dic['datasets'] = list()
    s, e = datas[0].get("order"), datas[-1].get("order")
    with codecs.open(output_path, "w", encoding="utf-8") as f:
        for data in datas:
            order = data.get("order")
            content = data.get("content").replace("\n", " ").replace("\r\n", " ")
            json_dic['datasets'].append({
                "content": content,
                "isSave": False,
                # "order": switch_order(len(str(order)), order),
                "order": order,
                "transContent": content
            })
        json.dump(json_dic, f)
    return s, e


def write_2_xlsx(datas, output_path):
    s, e = datas[0].get("order"), datas[-1].get("order")
    wb = Workbook()
    ws = wb.create_sheet("Sheet1")
    index = 1
    for data in datas:
        order = data.get("order")
        content = data.get("content").replace("\n", " ").replace("\r\n", " ")
        ws.cell(row=index, column=1, value=content)
        ws.cell(row=index, column=2, value=order)
        index += 1
    wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
    wb.save(output_path)
    return s, e


def update_content_status(col_name, use, s, e):
    col = mongodb[col_name]
    ret = col.update_many({"order": {"$gte": s, "$lte": e}}, {'$set': {'handleStatus': 1, "use": use}})
    print("成功更新了 %s 条数据的使用状态" % ret.matched_count)


def choose_colName_and_datasetName(language):
    names = {
        "荷兰": "col_dutch_text",
        "瑞典": "col_swedish_text",
        "丹麦": "col_danish_text",
        "泰米尔": "col_tamil_text",
        "老挝": "col_laotian_text",
        "波斯": "col_farsi_text",
        "挪威": "col_norwegian_text",
        "尼泊尔": "col_nepali_text",
        "爪哇": "col_javanese_text",
        "马拉地": "col_marathi_text",
        "巽他": "col_sundanese_text",
        "泰卢固": "col_telugu_text",
        "乌尔都": "col_urdu_text",
        "新蒙": "col_new_mongolian_text",
        "土库曼": "col_turkmen_text",
        "普什图": "col_pashto_text",
        "豪萨": "col_hausa_text",
        "乌兹别克": "col_uzbek_text",
        "哈萨克": "col_kazakh_text",
        "塔吉克": "col_tajik_text"
    }
    res = names.get(language)
    if not res:
        print("未找到对应的语言预置信息，请检查输入语言类型是否正确")
        exit(1)
    return res


def main(language, format, num, single_json_count, output_path):
    print("--- script start ---")
    count = num * single_json_count
    col_name = choose_colName_and_datasetName(language)
    datas = get_col_data(col_name, count)
    print("当前数据库未使用文本还有>>> %s" % len(datas))
    if not datas:
        print("MongoDB中未获取到可用数据")
        exit(1)
    export_num = math.floor(len(datas)/single_json_count)
    for i in range(export_num):
        _datas = datas[i * single_json_count:(i+1)*single_json_count]
        # prefix = time.strftime('%Y%m%d%H')
        # _path = os.path.join(output_path, "%s_%s_%s_%s" % (prefix, language, len(_datas), i))
        s_order = switch_order(len(str(i+50)), i+50)  # 哈萨克语已导出49
        # s_order = switch_order(len(str(i+1)), i+1)  # 哈萨克语已导出49
        _path = os.path.join(output_path, f"Kazakhstan_{s_order}")
        if format == "1":
            use = 1
            s, e = write_2_txt(_datas, _path + ".txt")
        elif format == "2":
            use = 2
            s, e = write_2_json(_datas, _path + ".json")
        elif format == "3":
            use = 3
            s, e = write_2_xlsx(_datas, _path + ".xlsx")
        else:
            use = 4
            _, _ = write_2_txt(_datas, _path + ".txt")
            s, e = write_2_json(_datas, _path + ".json")
        print(s, e)
        update_content_status(col_name, use, s, e)
    print("--- script end ---")


if __name__ == '__main__':
    language = input("请输入需要导出的语言，如丹麦、荷兰等：")
    format = input("请输入需要导出的格式, 1代表TXT，2代表json，3代表Excel，默认和其它任意键为TXT和JSON：")
    format = format if format == "1" or format == "2" or format == "3" else ""
    single_json_count = input("请输入需要导出单个文本句数，默认1100：")
    single_json_count = int(single_json_count) if single_json_count else 1100
    num = int(input("请输入需要导出的文件数量："))
    output_path = input("请输入导出文件的存放目录：")
    if not all((language, num, output_path)):
        print("输入参数不完整，请重新输入")
        exit(1)
    main(language, format, num, single_json_count, output_path)