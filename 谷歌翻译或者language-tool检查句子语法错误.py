# !/usr/bin/env python3
# -*- coding: utf-8 -*-

# @Time    : 2019/11/19 17:26
# @File    : 谷歌翻译检测语法错误.py
# @Date    : 2019/11/19
# @Author  : Yuwenjun
# @Desc    : python调用谷歌翻译API，过滤掉语法错误的句子
import json
import os
import random
import openpyxl
import requests
import execjs
import argparse
import language_check
import time
from pathlib import Path
from lxml import etree


class Py4Js(object):
    """生成tk值，来源于网络,感谢大神"""
    def __init__(self):
        self.ctx = execjs.compile(""" 
        function getTk(a) {
            bo = function(a) {
                return function() {
                    return a
                }
            };
            co = function(a, b) {
            for (var c = 0; c < b.length - 2; c += 3) {
                var d = b.charAt(c + 2);
                d = "a" <= d ? d.charCodeAt(0) - 87 : Number(d);
                d = "+" == b.charAt(c + 1) ? a >>> d : a << d;
                a = "+" == b.charAt(c) ? a + d & 4294967295 : a ^ d
            }
            return a
            };
            eo = '437307.4239319354';
            if (null !== eo)
                var b = eo;
            else {
                b = bo(String.fromCharCode(84));
                var c = bo(String.fromCharCode(75));
                b = [b(), b()];
                b[1] = c();
                b = (eo = window[b.join(c())] || "") || ""
            }
            var d = bo(String.fromCharCode(116));
            c = bo(String.fromCharCode(107));
            d = [d(), d()];
            d[1] = c();
            c = "&" + d.join("") + "=";
            d = b.split(".");
            b = Number(d[0]) || 0;
            for (var e = [], f = 0, g = 0; g < a.length; g++) {
                var h = a.charCodeAt(g);
                128 > h ? e[f++] = h : (2048 > h ? e[f++] = h >> 6 | 192 : (55296 == (h & 64512) && g + 1 < a.length && 56320 == (a.charCodeAt(g + 1) & 64512) ? (h = 65536 + ((h & 1023) << 10) + (a.charCodeAt(++g) & 1023),
                e[f++] = h >> 18 | 240,
                e[f++] = h >> 12 & 63 | 128) : e[f++] = h >> 12 | 224,
                e[f++] = h >> 6 & 63 | 128),
                e[f++] = h & 63 | 128)
            }
            a = b;
            for (f = 0; f < e.length; f++)
                a += e[f],
                a = co(a, "+-a^+6");
            a = co(a, "+-3^+b+-f");
            a ^= Number(d[1]) || 0;
            0 > a && (a = (a & 2147483647) + 2147483648);
            a %= 1E6;
            return a.toString() + "." + (a ^ b)
        }
    """)

    def getTk(self, text):
        """google translate请求参数中tk值是根据内容实时变化的，是由js动态生成，因此此函数调用js代码执行后返回text对应的tk值"""
        return self.ctx.call("getTk", text)


class CheckSentencesError(object):
    def __init__(self, language, checker, excel_path, col_index, head, sheet_name):
        self.headers = {
            'authority': 'translate.google.cn',
            'path': '',
            'method': 'GET',
            'scheme': 'https',
            'accept': '*/*',
            'accept-encoding': 'gzip, deflate, br',
            'accept-language': 'zh-CN,zh;q=0.9',
            'cookie': '',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64)  AppleWebKit/537.36 (KHTML, like Gecko) '
                          'Chrome/63.0.3239.108 Safari/537.36',
            'x-client-data': 'CIu2yQEIpbbJAQjBtskBCKmdygEI4qjKAQjOsMoBCOqxygEI97TKARirpMoB'
        }
        # self.proxy_url = "https://www.xicidaili.com/wn/1"
        # self.proxy_list = list()
        self.language = language
        self.excel_path = excel_path
        self.checker = checker
        self.col_index = col_index
        self.head = head
        self.sheet_name = sheet_name

    def get_language_code(self):
        iso_lan_dic = {
            "nl": "荷兰",
            "sv": "瑞典",
            "da": "丹麦",
            "ta": "泰米尔",
            "lo": "老挝",
            "fa": "波斯",
            "no": "挪威",
            "ne": "尼泊尔",
            "jv": "爪哇",
            "mr": "马拉地",
            "su": "巽他语",
            "te": "泰卢固",
            "ur": "乌尔都"
        }
        return iso_lan_dic.get(self.language)

    def parse_url(self, url):
        # response = requests.get(url, headers=self.headers, proxies=random.choice(self.proxy_list), timeout=5)
        response = requests.get(url, headers=self.headers, timeout=5)
        return response.content.decode()

    def bulid_proxy_ip_pool(self):
        response = requests.get(url=self.proxy_url, headers=self.headers)
        etree_obj = etree.HTML(response.text)
        ip_list = etree_obj.xpath("//tr[@class='odd']")
        item = []
        for ip in ip_list:
            ip_num = ip.xpath('./td[2]/text()')[0]
            port_num = ip.xpath('./td[3]/text()')[0]
            https = ip_num + ':' + port_num
            item.append(https)
        for it in item:
            try:
                proxy = {
                    'https': it
                }
                url_test = "https://translate.google.cn"
                res = requests.get(url=url_test, proxies=proxy, headers=self.headers, timeout=0.5)
                # 打印检测信息，elapsed.total_seconds()获取响应的时间
                print(it + '--', res.elapsed.total_seconds())
                self.proxy_list.append(proxy)
            except Exception as e:
                print(e)

    def build_url(self, text, tk, sl):
        baseUrl = 'https://translate.google.cn/translate_a/single'
        baseUrl += '?client=t&'
        baseUrl += 'sl={}&'.format(sl)
        baseUrl += 'tl=zh-CN&'
        baseUrl += 'hl=zh-CN&'
        baseUrl += 'dt=at&'
        baseUrl += 'dt=bd&'
        baseUrl += 'dt=ex&'
        baseUrl += 'dt=ld&'
        baseUrl += 'dt=md&'
        baseUrl += 'dt=qca&'
        baseUrl += 'dt=rw&'
        baseUrl += 'dt=rm&'
        baseUrl += 'dt=ss&'
        baseUrl += 'dt=t&'
        baseUrl += 'clearbtn=1&'
        # baseUrl += 'ie=UTF-8&&'
        # baseUrl += 'oe=UTF-8&'
        baseUrl += 'otf=1&'
        baseUrl += 'pc=1&'
        baseUrl += 'ssel=3&'
        baseUrl += 'tsel=3&'
        baseUrl += 'kc=2&'
        baseUrl += 'tk=' + str(tk) + '&'
        baseUrl += 'q=' + text
        return baseUrl

    def google_translate(self, text, tk, sl):
        url = self.build_url(text, tk, sl)
        print(url)
        res = True
        try:
            ret = self.parse_url(url)
            result = json.loads(ret)
            if result[7] is not None:
                # 如果我们文本输错，提示你是不是要找xxx的话，那么重新把xxx正确的翻译之后返回
                # 谷歌返回的结果是一个json格式的数据，我们将其变成一个嵌套的list，可以发现该list长度为9，第零个元素就是翻译结果,第七个结果是一些提示信息。
                # correct_text = result[7][0].replace('<b><i>', ' ').replace('</i></b>', '')
                # print(correct_text)
                # correct_url = self.build_url(correct_text, tk, sl)
                # correct_ret = self.parse_url(correct_url)
                # new_result = json.loads(correct_ret)
                # res = new_result[0][0][0]
                res = False
            # else:
            # res = result[0][0][0]
        except Exception as e:
            print("翻译" + text + "失败")
            print("错误信息: %s", e)
        finally:
            return res

    def get_language_tool_obj(self, iso_code):
        if iso_code in ["auto", "ast", "be", "br", "ca", "nl", "en-US", "eo", "fr", "gl", "de-DE", "el", "it", "ja",
                        "km", "fa", "pl", "pt", "ro", "ru", "zh", "sk", "sl", "es", "sv", "tl", "ta", "uk"]:
            tool = language_check.LanguageTool(iso_code)
            return tool
        else:
            print("不支持改语言类型")
            exit(1)

    def language_checker(self, tool, text):
        res = True
        try:
            matches = tool.check(text)
            if matches:
                res = False
        except Exception as e:
            print("检查" + text + "失败")
            print("错误信息: %s", e)
        finally:
            return res

    def run(self):
        js = Py4Js()
        # sl是要翻译的目标语种，值参照ISO 639-1标准，如果翻译成中文"zh/zh-CN简体中文"
        sl = self.language if self.get_language_code() else "auto"

        tool = self.get_language_tool_obj(sl) if self.checker == "language_tool" else None

        # 构建IP代理池
        # self.bulid_proxy_ip_pool()
        # print(self.proxy_list)

        # 以可编辑模式读取Excel
        wb = openpyxl.load_workbook(filename=self.excel_path)
        wb_new = openpyxl.Workbook()
        sheet_names = list()
        # 判断Excel文件是处理单一sheet还是所有表
        if self.sheet_name is not None:
            sheet_names.append(self.sheet_name)
        else:
            sheet_names.extend(wb.sheetnames)
        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            ws_new = wb_new.create_sheet("new_" + sheet_name)
            maxRows = ws.max_row
            maxCloumn = ws.max_column
            if self.col_index > maxCloumn:
                print("Excel文件sheet>>>{}输入的文本所在列数大于该sheet最大列数".format(sheet_name))
                continue
            print("start checking Sheet>>> %s - total text count % d" % (sheet_name, maxRows))
            count = 1
            for row in ws.iter_rows(min_row=self.head, max_col=maxCloumn, max_row=maxRows):
                check_text = row[self.col_index-1].value
                if check_text is None:
                    continue
                else:
                    check_text = check_text.strip()
                    tk = js.getTk(check_text)
                    if tool:
                        check_ret = self.language_checker(tool, check_text)
                    else:
                        check_ret = self.google_translate(check_text, tk, sl)
                    print("total {} now {} - checking {}".format(maxRows, count, check_text))
                    if check_ret is False:
                        # row_index = row[self.col_index-1].row
                        # ws.delete_rows(row_index)
                        print("⚠ language check not pass: %s" % check_text)
                    else:
                        print("✔ language check pass: %s" % check_text)
                        ws_new.append([cell.value for cell in row])
                count += 1
        p = Path(self.excel_path)
        new_excel_path = os.path.join(p.parent, "new_" + p.name)
        wb_new.remove_sheet(wb_new.get_sheet_by_name("Sheet"))
        wb_new.save(new_excel_path)


if __name__ == '__main__':
    ap = argparse.ArgumentParser(usage='%(prog)s [options] -p excel_path', description='pluck sentences')
    ap.add_argument("-l", "--language", type=str, default=None,
                    help="language to check for grammatical errors")
    ap.add_argument("-p", "--excel_path", type=str, required=True,
                    help="path to input Excel")
    ap.add_argument("-c", "--col_index", type=int, default=1,
                    help="column index that needs to check for text syntax errors, column A corresponds to 1")
    ap.add_argument("-head", "--head", type=int, default=1,
                    help="nearest multiple of 32 for resized width")
    ap.add_argument("-checker", "--checker", type=str, default="google",
                    help="nearest multiple of 32 for resized width")
    ap.add_argument("-s", "--sheet_name", type=str, default=None,
                    help="special sheet_name to check for grammatical errors")
    ap.print_help()

    args = vars(ap.parse_args())
    main = CheckSentencesError(**args).run()
    # main = CheckSentencesError(language="da", excel_path=r"F:\test1\丹麦语\丹麦语.xlsx", col_index=1, head=1, sheet_name='ekkofilm介绍').run()
    #python D:\projects\study\spider\谷歌翻译或者language-tool检查句子语法错误.py  -p "C:\Users\chenye\Desktop\lan\1115-瑞典电子书\10000分句\1115-瑞典电子书-副本_sheet1_0 - 副本.xlsx" -l sv -checker language_tool
    #python D:\projects\study\spider\谷歌翻译或者language-tool检查句子语法错误.py  -p "E:\dataset\北欧-我的文本\已交付\20191114-挪威-14万-新闻-养生-abcnyheter.no.xlsx" -l sv -checker language_tool