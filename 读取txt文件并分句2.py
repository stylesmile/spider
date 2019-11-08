# !/usr/bin/env python3
# -*- coding: utf-8 -*-

# @Time    : 2019/10/28 11:22
# @File    : 读取TSV文件并分句.py
# @Date    : 2019/10/28
# @Author  : Yuwenjun
# @Desc    :

from openpyxl import Workbook
import codecs
import re


letter_regex = re.compile(r'[a-zA-Z]')
# 加减乘除
computer_symbol = re.compile(r"[＋－×÷﹢﹣±＝=]")
# 去掉这些符号
special_symbol = re.compile(r"[©@#'，。★、【】《》？“”‘’！[\]_`{|\u4e00-\u9fa5}~]+")
# 数字
num_regex = re.compile(r'[0-9]')
# 特殊字符 放在另外一个文本
kuohao_regex = re.compile(r"[<>()?$;]")


def clean_tsv_content(contents):
    """
    读取tsv文件并分句
    :param 文本内容
    :return: list
    """
    data = list()
    for content in contents:
        e = 0
        for k, v in enumerate(content):
            if k == 0:
                if v == "۔":
                    e = 1
            elif k < len(content) - 1:  # 分句 根据 .?!;分句
                # 句号
                if v == ".":
                    if content[k - 1] != "." and content[k + 1] != ".":
                        if content[k - 1].isdecimal() and content[k + 1].isdecimal():
                            continue
                        data.append(content[e:k + 1].strip())
                        e = k + 1
                # 问号
                if v == "?":
                    if content[k - 1] != "?" and content[k + 1] != "?" and content[k + 1] != ")":
                        data.append(content[e:k + 1].strip())
                        e = k + 1
                if v == "!":
                    if content[k - 1] != "!" and content[k + 1] != ")":
                        data.append(content[e:k + 1].strip())
                        e = k + 1
                if v == ";":
                    data.append(content[e:k + 1].strip())
                    e = k + 1
                # if v == "|":
                #     data.append(content[e:k + 1].strip())
                #     e = k + 1
            else:
                data.append(content[e:].strip())
    return data


def main(input_path, excel_outpath):
    """
    将TXT清洗后的内容，断句写入Excel
    :param tsv文件路径
    :return:
    """
    wb = Workbook()
    # 筛选的句子保存到2个sheet
    ws1 = wb.create_sheet("去符号数字的句子")
    ws2 = wb.create_sheet("需手动筛选的句子")
    index = 1
    index2 = 1
    #with codecs.open(input_path, "r", encoding="utf-8-sig") as f:
    with codecs.open(input_path, "rb") as f:
        contents = f.readlines()
        datas = clean_tsv_content(contents)
        for data in datas:
            data = data.decode('utf-8')
            if data is None or len(data) < 15:
                continue
            if data.count("،") >= 2:
                continue
            if computer_symbol.findall(data):
                continue
            if num_regex.findall(data) or kuohao_regex.findall(data):
                ws2.cell(row=index2, column=1, value=data)
                index2 += 1
                continue
            ws1.cell(row=index, column=1, value=data)
            index += 1
    wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
    wb.save(excel_outpath)


if __name__ == '__main__':
    #txt_path = input("请输入txt文本路径：")
    #excel_path = input("请输入excel保存地址：")
    # 请输入txt文本路径
    # txt_path = r"E:\dataset\E-BooksNL-[334]NIEUWSEPTEMBER2019\2Amerge.txt"
    # 请输入excel保存地址
    # excel_path = r"E:\dataset\E-BooksNL-[334]NIEUWSEPTEMBER2019\2Amerge.xlsx"

    # txt_path = r"E:\dataset\瑞典10-30-20191107.txt"
    # excel_path = r"E:\dataset\瑞典10-30-20191107.txt.xlsx"

    # txt_path = r"E:\dataset\瑞典科技新闻-20191107.txt"
    # excel_path = r"E:\dataset\瑞典科技新闻-20191107.txt.xlsx"

    #txt_path = r"D:\360极速浏览器下载\se\merge.txt"
    #excel_path = r"D:\360极速浏览器下载\se\merge.txt.xlsx"

    # txt_path = r"E:\dataset\E-BooksNL-[334]NIEUWSEPTEMBER2019\A2\1merge.txt"
    # excel_path = r"E:\dataset\E-BooksNL-[334]NIEUWSEPTEMBER2019\A2\1merge.txt.xlsx"

    # txt_path = r"E:\dataset\\挪威语vildmedkrimi网络评论-20191107.txt"
    # excel_path = r"E:\dataset\\挪威语vildmedkrimi网络评论-20191107.txt.xlsx"
    # txt_path = r"E:\dataset\挪威语——新闻——科技新闻-sv.imwu-nl.com2000-4000-20191107.txt"
    # excel_path = r"E:\dataset\挪威语——新闻——科技新闻-sv.imwu-nl.com2000-4000-20191107.txt.xlsx"

    txt_path = r"E:\dataset\挪威语vildmedkrimi网络评论-20191107-1.txt"
    excel_path = r"E:\dataset\挪威语vildmedkrimi网络评论-20191107-1.txt.xlsx"

    txt_path = r"E:\dataset\挪威语abcnyheter.no-1000-1108.txt"
    excel_path = r"E:\dataset\挪威语abcnyheter.no-1000-1108.txt.xlsx"

    main(txt_path, excel_path)
