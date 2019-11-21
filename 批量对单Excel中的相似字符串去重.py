# !/usr/bin/env python3
# -*- coding: utf-8 -*-

# @Time    : 2019/11/17 13:32
# @File    : 批量对单Excel中的相似字符串去重.py
# @Date    : 2019/11/17
# @Author  : Yuwenjun
# @Desc    : 去除标点符号后，比较字符串相似度,过滤掉相似度大于70%的


import openpyxl
import os
import difflib
import re

mark_symbol = re.compile(r"[,.:;?()[]&!*@#$%<>+-/'\"]")


def main(file_path, output_path, sheet_name):
    excel_names = [name for name in os.listdir(file_path) if name.endswith(".xlsx")]
    for excel_name in excel_names:
        excel_path = os.path.join(file_path, excel_name)
        wb_new = openpyxl.Workbook()
        ws_new = wb_new.create_sheet("")
        wb_old = openpyxl.load_workbook(filename=excel_path, read_only=True)
        if sheet_name == "":
            ws = wb_old.active
        else:
            ws = wb_old[sheet_name]
        maxRows = ws.max_row
        index = 1
        print("开始对Excel文件>>>{}进行相似度去重".format(excel_name))
        for row in ws.iter_rows(min_row=1, max_col=4, max_row=maxRows):
            text = row[0].value
            if index < 2:
                ws_new.cell(row=index, column=1, value=text)
                index += 1
                continue
            else:
                if not text:
                    continue
                # text_clean = mark_symbol.sub("", text)
                # print(text_clean)
                tag = False
                for row_new in ws_new.rows:
                    text2 = row_new[0].value
                    # text2_clean = mark_symbol.sub("", text2)
                    ret = difflib.SequenceMatcher(None, text, text2).quick_ratio()
                    if ret >= 0.95:
                        tag = True
                        break
                if not tag:
                    ws_new.cell(row=index, column=1, value=text)
                    index += 1
        new_excel_name = excel_name + "new"
        if output_path is None:
            excel_file_outpath = os.path.join(file_path, new_excel_name)
        else:
            excel_file_outpath = os.path.join(file_path, new_excel_name)
        wb_new.save(excel_file_outpath)


if __name__ == '__main__':
    file_path = input("请输入待处理的Excel文件夹路径，如f:\\test1>>>")
    output_path = input("请输入待输出的Excel文件夹路径，如不填写则去重后的Excel保存到原Excel路径中>>>")
    sheet_name = input("请输入含有文本的sheet名，默认为Mysheet>>>")
    if not os.path.exists(file_path):
        print("请输入有效的待处理Excel路径")
        exit(1)
    if not sheet_name:
        sheet_name = "Mysheet"
    main(file_path, output_path, sheet_name)
