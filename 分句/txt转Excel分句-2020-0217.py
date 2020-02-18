# !/usr/bin/env python3
# -*- coding: utf-8 -*-

# @Time    : 2019/11/11 13:55
# @File    : txt转Excel.py
# @Date    : 2019/11/11
# @Author  : Yuwenjun
# @Desc    : 将pdf转换后的TXT文件转换成Excel

import os
import codecs
import re
from openpyxl import Workbook


letter_regex = re.compile('[a-zA-Z]+')  # 英文大小写字母正则
computer_symbol = re.compile(r"[＋－×÷﹢﹣±＝=#《》]+")  # 常用数学计算符号正则
special_symbol = re.compile(r"[#'，。★、【】《》？“”‘’！[\]_`{|\u4e00-\u9fa5}~]+")  # 部分特殊符号正则
num_regex = re.compile('[0-9]+')  # 数字正则
kuohao_regex = re.compile(r"[<>():?$;؟]+")  # 问句、括号、冒号等语气可能不确定符号的正则
mark_symbol = re.compile('[,.:;?()[]<>&!#%"\'”“]+')  # 规则允许符号的正则
#replace_regex = re.compile(r"[😂【】]+")  # 需要被替换符号的正则
replace_regex = re.compile(r"[˝\*«»?()\[\]<>•…–-—\-—﹣😂【】]+")  # 需要被替换符号的正则 替换为空格


def clean_tsv_content(contents):
    """
    读取tsv文件并分句
    :param 文本内容
    :return: list
    """
    data = list()
    for content in contents:
        e = 0
        for k, v in enumerate(content):
            if k == 0:
                if v == ".":
                    e = 1
            elif k < len(content) - 1:
                if v == ".":
                    if content[k - 1].isdecimal() and content[k + 1].isdecimal():
                        continue
                    if content[k-2:k] == "bl" and content[k+1] == "a":
                        continue
                    data.append(content[e:k + 1].strip())
                    e = k + 1
                if v == "۔":
                    if content[k - 1] != "۔" and content[k + 1] != "۔":
                        if content[k - 1].isdecimal() and content[k + 1].isdecimal():
                            continue
                        data.append(content[e:k + 1].strip())
                        e = k + 1
                if v == "؟":
                    if content[k - 1] != "؟" and content[k + 1] != "؟" and content[k + 1] != ")":
                        data.append(content[e:k + 1].strip())
                        e = k + 1
                if v == "!":
                    if content[k + 1] != "!" and content[k + 1] != ")":
                        data.append(content[e:k + 1].strip())
                        e = k + 1
                if v == "?":
                    data.append(content[e:k + 1].strip())
                    e = k + 1
                if v == "।":
                    data.append(content[e:k + 1].strip())
                    e = k + 1
            else:
                data.append(content[e:].strip())
    return data


def merge_content(contents):
    content_list = list()
    tmp = ""
    for k, content in enumerate(contents):
        if k == len(contents) - 1:
            tmp += content.strip()
            content_list.append(tmp)
            tmp = ""
        else:
            if content.strip():
                if content.endswith((".", '."', "—", "-", ".'")):
                    tmp += content.strip()
                    content_list.append(tmp)
                    tmp = ""
                else:
                    if mark_symbol.findall(content.strip()[-1]):
                        tmp += content.strip()
                    else:
                        tmp += content.strip() + " "
    return content_list


def main(input_path, excel_outpath):
    """
    将TXT清洗后的内容，断句写入Excel
    :param input_path: txt存放文件夹目录
    :param excel_outpath: Excel保存地址
    :return:
    """
    wb = Workbook()  # 实例化一个工作簿
    ws1 = wb.create_sheet("Sheet1")  # 创建sheet1
    ws2 = wb.create_sheet("需手动筛选的句子")  # 创建另一个sheet
    index = 1  # index和index1都是写入时Excel的行标
    index2 = 1
    # 获取输入文件夹下的所有TXT名
    txt_names = [name for name in os.listdir(input_path) if name.endswith(".txt")]
    for txt_name in txt_names:  # 遍历处理每个TXT
        print(txt_name)
        txt_path = os.path.join(input_path, txt_name)  # 拼接路径
        with codecs.open(txt_path, "r", encoding="utf-8", errors="ignore") as f:  # 读取TXT
            contents = merge_content(f.readlines())  # 文本按行找符号合并
            datas = clean_tsv_content(contents)  # 分局
            for data in datas:
                if data is None or len(data) < 10:  # 过滤过短的文本
                    continue
                if computer_symbol.findall(data):  # 过滤含有数学计算符号的文本
                    continue
                # if letter_regex.findall(data):  # 过滤英文文本
                #     continue
                if data[0] == "," or data[0] == ".":  # 对于文本首字母为逗号句号的切片
                    data = data[1:]
                if data[0] == "[":  # 过滤文本首字母为左中括号的，一般包含它的句子不完整
                    continue
                # 对于有数字和不确定语气符号的句子放到另一个sheet中
                if num_regex.findall(data) or kuohao_regex.findall(data) or special_symbol.findall(data):
                    ws2.cell(row=index2, column=1, value=data.replace("'", "").replace('"', ""))
                    ws2.cell(row=index2, column=2, value=txt_name)
                    index2 += 1
                    continue
                text = data.replace("'", "").replace('"', "")  # 替换掉文本中的引号
                text = replace_regex.sub(" ", text)  # 替换文本中的需替换符号为空格
                ws1.cell(row=index, column=1, value=text)
                ws1.cell(row=index, column=2, value=txt_name)
                index += 1
    wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
    wb.save(excel_outpath)


if __name__ == '__main__':
    #txt_path = input("请输入txt文本路径：")
    #excel_path = input("请输入excel保存地址：")
    txt_path = r"F:\data\hansa2020-0217"
    excel_path = r"F:\data\hansa2020-0217\hansa2020-0217.xlsx"
    main(txt_path, excel_path)

# 去重 去–- 双空格替换空格
# =trim(A1)
# =LEN(TRIM(C1))-LEN(SUBSTITUTE(TRIM(C1)," ",))+1
# =LEFT(C1,1)
# =EXACT(UPPER(E1),E1)
# 去重