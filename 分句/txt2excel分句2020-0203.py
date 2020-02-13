# !/usr/bin/env python3
# -*- coding: utf-8 -*-

# @Time    : 2019/11/11 13:55
# @File    : txt转Excel.py
# @Date    : 2019/11/11
# @Author  : Yuwenjun
# @Desc    : 将pdf转换后的TXT文件转换成Excel

import os
import codecs
import re
from openpyxl import Workbook

letter_regex = re.compile('[a-zA-Z]')
#
#computer_symbol = re.compile(r"[＋－×÷﹢﹣±＝=#《》]")
computer_symbol = re.compile(r"[＋×÷﹢＝=#《》]")
"""
符号说明
« »挪威书名号
用空替換這些符號 –-
单独提出来，放一个是sheet
"""
# special_symbol = re.compile(r"[#'，。★、【】《》？“”‘’！[\]_`{|\u4e00-\u9fa5}~]+")
special_symbol = re.compile(r"[•…\/\*\{\}\[\]\»“”※§©!:;«»+@#'，。★、【】《》？！[\]_`{|\u4e00-\u9fa5}~]")
num_regex = re.compile('[0-9]')
# 有符号去掉
kuohao_regex = re.compile(r"[<>():?$;؟]")
# 用空格替換這些符號
#mark_symbol = re.compile('[-—,.:;?()[]<>&!#%"\'”“]')
mark_symbol = re.compile('[–-——﹣\-\–,.:;?()[]<>&!#%"\‘\’\'”“]')


def clean_tsv_content(contents):
    """
    读取tsv文件并分句
    :param 文本内容
    :return: list
    """
    data = list()
    for content in contents:
        e = 0
        for k, v in enumerate(content):
            if k == 0:
                if v == ".":
                    e = 1
            elif k < len(content) - 1:
                if v == ".":
                    if content[k - 1].isdecimal() and content[k + 1].isdecimal():
                        continue
                    if content[k - 2:k] == "bl" and content[k + 1] == "a":
                        continue
                    data.append(content[e:k + 1].strip())
                    e = k + 1
                if v == "۔":
                    if content[k - 1] != "۔" and content[k + 1] != "۔":
                        if content[k - 1].isdecimal() and content[k + 1].isdecimal():
                            continue
                        data.append(content[e:k + 1].strip())
                        e = k + 1
                if v == "؟":
                    if content[k - 1] != "؟" and content[k + 1] != "؟" and content[k + 1] != ")":
                        data.append(content[e:k + 1].strip())
                        e = k + 1
                if v == "!":
                    if content[k + 1] != "!" and content[k + 1] != ")":
                        data.append(content[e:k + 1].strip())
                        e = k + 1
                if v == "?":
                    data.append(content[e:k + 1].strip())
                    e = k + 1
                if v == "।":
                    data.append(content[e:k + 1].strip())
                    e = k + 1
            else:
                data.append(content[e:].strip())
    return data


def merge_content(contents):
    content_list = list()
    tmp = ""
    for k, content in enumerate(contents):
        if k == len(contents) - 1:
            tmp += content.strip()
            content_list.append(tmp)
            tmp = ""
        else:
            if content.strip():
                if content.endswith((".", '."', "—", "-", ".'")):
                    tmp += content.strip()
                    content_list.append(tmp)
                    tmp = ""
                else:
                    if mark_symbol.findall(content.strip()[-1]):
                        tmp += content.strip()
                    else:
                        tmp += content.strip() + " "
    return content_list


def main(input_path, excel_outpath):
    """
    将TXT清洗后的内容，断句写入Excel
    :param tsv文件路径
    :return:
    """
    wb = Workbook()
    ws1 = wb.create_sheet("Sheet1")
    ws2 = wb.create_sheet("需手动筛选的句子")
    index = 1
    index2 = 1
    txt_names = [name for name in os.listdir(input_path) if name.endswith(".txt")]
    for txt_name in txt_names:
        print(txt_name)
        txt_path = os.path.join(input_path, txt_name)
        with codecs.open(txt_path, "r", encoding="utf-8-sig") as f:
            contents = merge_content(f.readlines())
            datas = clean_tsv_content(contents)
            for data in datas:
                if data is None or len(data) < 10:
                    continue
                if computer_symbol.findall(data):
                    continue
                # if letter_regex.findall(data):
                #     continue
                if data[0] == "," or data[0] == ".":
                    data = data[1:]
                if data[0] == "[":
                    continue
                if num_regex.findall(data) or kuohao_regex.findall(data) or special_symbol.findall(data):
                    ws2.cell(row=index2, column=1, value=data.replace("'", "").replace('"', ""))
                    ws2.cell(row=index2, column=2, value=txt_name)
                    index2 += 1
                    continue
                ws1.cell(row=index, column=1, value=data.replace("'", "").replace('"', ""))
                ws1.cell(row=index, column=2, value=txt_name)
                index += 1
    wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
    wb.save(excel_outpath)


if __name__ == '__main__':
    # txt_path = input("请输入txt文本路径：")
    # excel_path = input("请输入excel保存地址：")

    txt_path = r"C:\Users\chenye\Desktop\data\no2020-0204"
    excel_path = r"C:\Users\chenye\Desktop\data\no2020-0204\no20200204.xlsx"

    txt_path = r"C:\Users\chenye\Desktop\data\no2020-0205"
    excel_path = r"C:\Users\chenye\Desktop\data\no2020-0205\no20200205.xlsx"

    txt_path = r"C:\Users\chenye\Desktop\data\no-ebook-2020-0205"
    excel_path = r"C:\Users\chenye\Desktop\data\no-ebook-2020-0205\no-ebook-2020-0205.xlsx"

    txt_path = r"C:\Users\chenye\Desktop\data\SW-2-5"
    excel_path = r"C:\Users\chenye\Desktop\data\SW-2-5\SW-2-5.xlsx"
    txt_path = r"C:\Users\chenye\Desktop\data\陈烨"
    excel_path = r"C:\Users\chenye\Desktop\data\陈烨\陈烨-2-5.xlsx"

    txt_path = r"F:\data\mn-2020-02-11"
    excel_path = r"F:\data\mn-2020-02-11\mn-2020-02-11.xlsx"
    main(txt_path, excel_path)

# 去重 去–- 双空格替换空格
# =trim(A1)
# =LEN(TRIM(C1))-LEN(SUBSTITUTE(TRIM(C1)," ",))+1
# =LEFT(C1,1)
# =EXACT(UPPER(E1),E1)
# 去重

# 去重 去–- 双空格替换空格
# =trim(A1)
# =LEN(TRIM(C1))-LEN(SUBSTITUTE(TRIM(C1)," ",))+1
# =LEFT(C1,1)
# =EXACT(UPPER(E1),E1)
# 去重